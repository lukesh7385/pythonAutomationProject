import inspect
import logging
import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_row)


def getColumnCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return (sheet.max_column)


def readData(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rownum, columnno).value


def writeData(file, sheetName, rownum, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(rownum, columnno).value = data
    workbook.save(file)


def fillGreenColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    greenFill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum, columnno).fill = greenFill
    workbook.save(file)


def fillRedColor(file, sheetName, rownum, columnno):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    redFill = PatternFill(start_color='ff0000',
                          end_color='ff0000',
                          fill_type='solid')
    sheet.cell(rownum, columnno).fill = redFill
    workbook.save(file)



# def getRowCount(file, sheetname):
#     Book = openpyxl.load_workbook(file)
#     Sheet = Book[sheetname]
#     return Sheet.max_row
#
#
# def readData(file, sheetname, rownum, colnum):
#     Book = openpyxl.load_workbook(file)
#     Sheet = Book[sheetname]
#     return Sheet.cell(row=rownum, column=colnum).value
#
#
# def writeData(file, sheetname, rownum, colnum, data):
#     Book = openpyxl.load_workbook(file)
#     Sheet = Book[sheetname]
#     Sheet.cell(row=rownum, column=colnum).value = data
#     Book.save(file)